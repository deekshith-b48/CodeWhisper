import os
import io
import csv
import json
import logging
from typing import Dict, Any, Optional, Tuple
from datetime import datetime
import PyPDF2
import docx
import openpyxl
import pandas as pd
from pathlib import Path

logger = logging.getLogger(__name__)

class FileProcessingService:
    def __init__(self):
        self.supported_extensions = {
            # Text files
            'txt': self._process_text_file,
            'md': self._process_text_file,
            'py': self._process_text_file,
            'js': self._process_text_file,
            'ts': self._process_text_file,
            'jsx': self._process_text_file,
            'tsx': self._process_text_file,
            'java': self._process_text_file,
            'cpp': self._process_text_file,
            'c': self._process_text_file,
            'h': self._process_text_file,
            'sql': self._process_text_file,
            'html': self._process_text_file,
            'htm': self._process_text_file,
            'xml': self._process_text_file,
            'json': self._process_text_file,
            'yaml': self._process_text_file,
            'yml': self._process_text_file,
            'sh': self._process_text_file,
            'dockerfile': self._process_text_file,
            
            # Data files
            'csv': self._process_csv_file,
            'xlsx': self._process_excel_file,
            'xls': self._process_excel_file,
            
            # Document files
            'pdf': self._process_pdf_file,
            'docx': self._process_docx_file,
            'doc': self._process_doc_file,
        }
    
    def can_process_file(self, filename: str) -> bool:
        """Check if the file can be processed"""
        extension = self._get_file_extension(filename)
        return extension in self.supported_extensions
    
    def get_supported_extensions(self) -> list:
        """Get list of supported file extensions"""
        return list(self.supported_extensions.keys())
    
    def process_file(self, file_content: bytes, filename: str, metadata: Dict[str, Any] = None) -> Tuple[str, Dict[str, Any]]:
        """
        Process a file and return its text content and metadata
        """
        try:
            extension = self._get_file_extension(filename)
            
            if extension not in self.supported_extensions:
                raise ValueError(f"Unsupported file type: {extension}")
            
            # Get the processing function
            process_func = self.supported_extensions[extension]
            
            # Process the file
            content, file_metadata = process_func(file_content, filename)
            
            # Merge with provided metadata
            if metadata:
                file_metadata.update(metadata)
            
            # Add processing metadata
            file_metadata.update({
                'processed_at': datetime.utcnow().isoformat(),
                'file_extension': extension,
                'original_filename': filename,
                'content_length': len(content),
                'processing_method': process_func.__name__
            })
            
            return content, file_metadata
            
        except Exception as e:
            logger.error(f"Error processing file {filename}: {str(e)}")
            raise
    
    def _get_file_extension(self, filename: str) -> str:
        """Extract file extension from filename"""
        return filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
    
    def _process_text_file(self, file_content: bytes, filename: str) -> Tuple[str, Dict[str, Any]]:
        """Process text-based files"""
        try:
            content = file_content.decode('utf-8')
            metadata = {
                'file_type': 'text',
                'encoding': 'utf-8',
                'language': self._detect_language(filename)
            }
            return content, metadata
        except UnicodeDecodeError:
            # Try other encodings
            for encoding in ['latin-1', 'cp1252', 'iso-8859-1']:
                try:
                    content = file_content.decode(encoding)
                    metadata = {
                        'file_type': 'text',
                        'encoding': encoding,
                        'language': self._detect_language(filename)
                    }
                    return content, metadata
                except UnicodeDecodeError:
                    continue
            raise ValueError(f"Could not decode file {filename} with any supported encoding")
    
    def _process_csv_file(self, file_content: bytes, filename: str) -> Tuple[str, Dict[str, Any]]:
        """Process CSV files"""
        try:
            # Try to decode as UTF-8 first
            csv_text = file_content.decode('utf-8')
        except UnicodeDecodeError:
            # Try other encodings
            for encoding in ['latin-1', 'cp1252', 'iso-8859-1']:
                try:
                    csv_text = file_content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise ValueError(f"Could not decode CSV file {filename}")
        
        # Parse CSV and convert to readable format
        csv_reader = csv.reader(io.StringIO(csv_text))
        rows = list(csv_reader)
        
        if not rows:
            return "", {'file_type': 'csv', 'rows': 0, 'columns': 0}
        
        # Convert to markdown table format
        content_lines = []
        for i, row in enumerate(rows):
            if i == 0:  # Header row
                content_lines.append("| " + " | ".join(str(cell) for cell in row) + " |")
                content_lines.append("| " + " | ".join(["---"] * len(row)) + " |")
            else:
                content_lines.append("| " + " | ".join(str(cell) for cell in row) + " |")
        
        content = "\n".join(content_lines)
        
        metadata = {
            'file_type': 'csv',
            'rows': len(rows),
            'columns': len(rows[0]) if rows else 0,
            'has_header': True
        }
        
        return content, metadata
    
    def _process_excel_file(self, file_content: bytes, filename: str) -> Tuple[str, Dict[str, Any]]:
        """Process Excel files (XLSX, XLS)"""
        try:
            # Load Excel file
            excel_file = io.BytesIO(file_content)
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            
            content_lines = []
            sheet_count = 0
            total_rows = 0
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_count += 1
                
                # Add sheet header
                content_lines.append(f"\n## Sheet: {sheet_name}\n")
                
                # Get all data from sheet
                data = []
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        data.append([str(cell) if cell is not None else "" for cell in row])
                
                if data:
                    # Convert to markdown table
                    for i, row in enumerate(data):
                        if i == 0:  # Header row
                            content_lines.append("| " + " | ".join(cell for cell in row) + " |")
                            content_lines.append("| " + " | ".join(["---"] * len(row)) + " |")
                        else:
                            content_lines.append("| " + " | ".join(cell for cell in row) + " |")
                    
                    total_rows += len(data)
                    content_lines.append(f"\n*Sheet contains {len(data)} rows*\n")
            
            content = "\n".join(content_lines)
            
            metadata = {
                'file_type': 'excel',
                'sheets': sheet_count,
                'total_rows': total_rows,
                'sheet_names': workbook.sheetnames
            }
            
            return content, metadata
            
        except Exception as e:
            logger.error(f"Error processing Excel file {filename}: {str(e)}")
            raise ValueError(f"Could not process Excel file: {str(e)}")
    
    def _process_pdf_file(self, file_content: bytes, filename: str) -> Tuple[str, Dict[str, Any]]:
        """Process PDF files"""
        try:
            pdf_file = io.BytesIO(file_content)
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            
            content_lines = []
            page_count = len(pdf_reader.pages)
            
            for page_num, page in enumerate(pdf_reader.pages, 1):
                try:
                    page_text = page.extract_text()
                    if page_text.strip():
                        content_lines.append(f"\n## Page {page_num}\n")
                        content_lines.append(page_text)
                except Exception as e:
                    logger.warning(f"Could not extract text from page {page_num}: {str(e)}")
                    content_lines.append(f"\n## Page {page_num}\n")
                    content_lines.append("*[Text extraction failed for this page]*")
            
            content = "\n".join(content_lines)
            
            metadata = {
                'file_type': 'pdf',
                'pages': page_count,
                'extraction_method': 'PyPDF2'
            }
            
            return content, metadata
            
        except Exception as e:
            logger.error(f"Error processing PDF file {filename}: {str(e)}")
            raise ValueError(f"Could not process PDF file: {str(e)}")
    
    def _process_docx_file(self, file_content: bytes, filename: str) -> Tuple[str, Dict[str, Any]]:
        """Process DOCX files"""
        try:
            docx_file = io.BytesIO(file_content)
            doc = docx.Document(docx_file)
            
            content_lines = []
            paragraph_count = 0
            
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    content_lines.append(paragraph.text)
                    paragraph_count += 1
            
            content = "\n\n".join(content_lines)
            
            metadata = {
                'file_type': 'docx',
                'paragraphs': paragraph_count,
                'extraction_method': 'python-docx'
            }
            
            return content, metadata
            
        except Exception as e:
            logger.error(f"Error processing DOCX file {filename}: {str(e)}")
            raise ValueError(f"Could not process DOCX file: {str(e)}")
    
    def _process_doc_file(self, file_content: bytes, filename: str) -> Tuple[str, Dict[str, Any]]:
        """Process DOC files (legacy Word format)"""
        try:
            # For DOC files, we'll try to convert them or provide a helpful message
            # This is a simplified approach - in production you might want to use
            # a more robust conversion library like antiword or catdoc
            
            content = f"""# Legacy Word Document: {filename}

This is a legacy Word document (.doc) format. For better processing, please convert this file to .docx format or provide the content as text.

**File Information:**
- Original filename: {filename}
- File size: {len(file_content)} bytes
- Format: Legacy Microsoft Word (.doc)

**Recommendation:** Convert this file to .docx format for better text extraction and processing."""
            
            metadata = {
                'file_type': 'doc',
                'legacy_format': True,
                'conversion_needed': True,
                'extraction_method': 'placeholder'
            }
            
            return content, metadata
            
        except Exception as e:
            logger.error(f"Error processing DOC file {filename}: {str(e)}")
            raise ValueError(f"Could not process DOC file: {str(e)}")
    
    def _detect_language(self, filename: str) -> str:
        """Detect programming language from filename"""
        extension = self._get_file_extension(filename)
        language_map = {
            'py': 'python',
            'js': 'javascript',
            'ts': 'typescript',
            'jsx': 'javascript',
            'tsx': 'typescript',
            'java': 'java',
            'cpp': 'cpp',
            'c': 'c',
            'h': 'c',
            'cs': 'csharp',
            'go': 'go',
            'rs': 'rust',
            'php': 'php',
            'rb': 'ruby',
            'swift': 'swift',
            'kt': 'kotlin',
            'scala': 'scala',
            'sql': 'sql',
            'html': 'html',
            'css': 'css',
            'scss': 'scss',
            'json': 'json',
            'xml': 'xml',
            'yaml': 'yaml',
            'yml': 'yaml',
            'md': 'markdown',
            'sh': 'bash',
            'dockerfile': 'dockerfile'
        }
        return language_map.get(extension, 'text') 